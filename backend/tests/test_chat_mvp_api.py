from collections.abc import Iterator
from concurrent.futures import ThreadPoolExecutor
from contextlib import contextmanager
from pathlib import Path
from threading import Barrier, Thread

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.adapters.dialogue import LangGraphChatWorkflow
from app.adapters.llm.fake_llm_client import FakeLlmClient
from app.adapters.repositories.sqlite_repository import SQLiteExcelAssetRepository
from app.adapters.storage.filesystem_storage import FilesystemExcelArtifactStorage
from app.adapters.workbook.openpyxl_reader import OpenpyxlWorkbookReader
from app.api.dependencies import (
    get_chat_service,
    get_current_user,
    get_document_summary_service,
    get_excel_asset_service,
    get_llm_preference_service,
    require_admin_user,
)
from app.application.chat.cancellation import (
    ChatCancellationRegistry,
    ChatCancellationToken,
    ChatRequestCancelledError,
)
from app.application.chat.service import ChatService
from app.application.document_summaries.service import DocumentSummaryService
from app.application.excel_assets.service import ExcelAssetService
from app.application.llm_preferences import WorkspaceLlmPreferenceService
from app.core.config import Settings
from app.core.errors import LlmRequestError
from app.domain.models import (
    AttachedDocument,
    ChatTurn,
    DocumentSummary,
    DraftAnswerBlock,
    DraftChatAnswer,
    DraftCitation,
    ExcelCitation,
    LlmPreference,
    SelectedDocument,
)
from app.main import app
from app.ports.chat_workflow import ChatWorkflow
from tests.auth_helpers import admin_user, member_user


@pytest.fixture
def client(tmp_path: Path) -> Iterator[TestClient]:
    repository = SQLiteExcelAssetRepository(tmp_path / "excel.sqlite3")
    excel_assets = ExcelAssetService(
        repository=repository,
        storage=FilesystemExcelArtifactStorage(tmp_path / "storage"),
        workbook_reader=OpenpyxlWorkbookReader(),
    )
    excel_assets.initialize()
    llm_client = FakeLlmClient()
    llm_preferences = WorkspaceLlmPreferenceService(
        repository=repository,
        settings=Settings(_env_file=None),
    )
    summaries = DocumentSummaryService(
        excel_assets=excel_assets,
        llm_client=llm_client,
        repository=repository,
        llm_preferences=llm_preferences,
    )
    chat = ChatService(
        excel_assets=excel_assets,
        summaries=summaries,
        llm_client=llm_client,
        sessions=repository,
        llm_preferences=llm_preferences,
    )
    app.dependency_overrides[get_excel_asset_service] = lambda: excel_assets
    app.dependency_overrides[get_document_summary_service] = lambda: summaries
    app.dependency_overrides[get_chat_service] = lambda: chat
    app.dependency_overrides[get_llm_preference_service] = lambda: llm_preferences
    app.dependency_overrides[get_current_user] = admin_user
    app.dependency_overrides[require_admin_user] = admin_user
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


def test_summary_generation_and_chat_framework_flow(
    client: TestClient,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "standards.xlsx"
    _write_xlsx_fixture(workbook_path)
    with workbook_path.open("rb") as workbook_file:
        upload_response = client.post(
            "/api/excel/files",
            files={
                "file": (
                    "standards.xlsx",
                    workbook_file,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert upload_response.status_code == 200
    version_id = upload_response.json()["version"]["version_id"]

    missing_summary_response = client.get(f"/api/excel/versions/{version_id}/summary")
    assert missing_summary_response.status_code == 404

    summary_response = client.post(f"/api/excel/versions/{version_id}/summary/generate")
    assert summary_response.status_code == 200
    assert summary_response.json()["version_id"] == version_id

    chat_response = client.post(
        "/api/excel/chat",
        json={"question": "What does the standards file contain?"},
    )

    assert chat_response.status_code == 200
    answer = chat_response.json()
    assert answer["selected_documents"][0]["version_id"] == version_id
    assert answer["answer_blocks"][0]["citation_ids"] == ["C1"]
    assert answer["citations"][0]["citation_id"] == "C1"
    assert (
        answer["citations"][0]["evidence_id"]
        == f"{version_id}::{answer['citations'][0]['sheet_id']}::S001_R1"
    )
    assert answer["citations"][0]["row_id"] == "S001_R1"
    assert answer["citations"][0]["row"][1:] == ["Code", "Date"]


def test_complete_user_workflow_with_fake_data_and_concurrent_reads(
    tmp_path: Path,
) -> None:
    with _client_with_llm(tmp_path, FakeLlmClient()) as client:
        workbook_path = tmp_path / "workflow.xlsx"
        _write_workflow_xlsx_fixture(workbook_path)

        with workbook_path.open("rb") as workbook_file:
            upload_response = client.post(
                "/api/excel/files",
                files={
                    "file": (
                        "workflow.xlsx",
                        workbook_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        assert upload_response.status_code == 200
        upload_payload = upload_response.json()
        file_id = upload_payload["file"]["file_id"]
        version_id = upload_payload["version"]["version_id"]
        first_sheet_id = upload_payload["sheets"][0]["sheet_id"]

        files_response = client.get("/api/excel/files")
        assert files_response.status_code == 200
        assert files_response.json()["files"][0]["file_id"] == file_id

        versions_response = client.get(f"/api/excel/files/{file_id}/versions")
        sheets_response = client.get(f"/api/excel/versions/{version_id}/sheets")
        preview_response = client.get(f"/api/excel/sheets/{first_sheet_id}/preview?limit=5")
        search_response = client.get(
            f"/api/excel/versions/{version_id}/search?query=Apex&limit=10"
        )
        assert versions_response.status_code == 200
        assert sheets_response.status_code == 200
        assert preview_response.status_code == 200
        assert preview_response.json()["rows"][1][1:] == ["Apex", "High", "North"]
        assert search_response.status_code == 200
        assert search_response.json()["total_matches"] == 1

        row_id = search_response.json()["matches"][0]["mapping"]["row_id"]
        lookup_response = client.get(f"/api/excel/sheets/{first_sheet_id}/rows/{row_id}")
        assert lookup_response.status_code == 200
        assert lookup_response.json()["row"][1:] == ["Apex", "High", "North"]

        summary_response = client.post(f"/api/excel/versions/{version_id}/summary/generate")
        assert summary_response.status_code == 200
        update_summary_response = client.patch(
            f"/api/excel/versions/{version_id}/summary",
            json={
                "document_title": "Workflow workbook",
                "summary_text": "Small workflow workbook for smoke testing.",
                "business_domain": "quality operations",
                "key_topics": ["Apex", "North"],
            },
        )
        assert update_summary_response.status_code == 200
        assert update_summary_response.json()["document_title"] == "Workflow workbook"

        session_response = client.post("/api/excel/chat/sessions")
        assert session_response.status_code == 200
        session_id = session_response.json()["session_id"]

        rename_response = client.patch(
            f"/api/excel/chat/sessions/{session_id}",
            json={"title": "Apex review"},
        )
        pin_response = client.patch(
            f"/api/excel/chat/sessions/{session_id}/pin",
            json={"pinned": True},
        )
        chat_response = client.post(
            f"/api/excel/chat/sessions/{session_id}/messages",
            json={"question": "What does Apex show?"},
        )
        assert rename_response.status_code == 200
        assert rename_response.json()["title"] == "Apex review"
        assert pin_response.status_code == 200
        assert pin_response.json()["pinned_at"] is not None
        assert chat_response.status_code == 200
        assert chat_response.json()["citations"]

        def fetch_preview_and_search(index: int) -> tuple[int, int]:
            preview = client.get(
                f"/api/excel/sheets/{first_sheet_id}/preview?offset={index % 2}&limit=2"
            )
            search = client.get(
                f"/api/excel/versions/{version_id}/search?query=North&limit=5"
            )
            return preview.status_code, search.status_code

        with ThreadPoolExecutor(max_workers=6) as executor:
            statuses = list(executor.map(fetch_preview_and_search, range(12)))
        assert statuses == [(200, 200)] * 12

        hide_response = client.patch(
            f"/api/excel/files/{file_id}/visibility",
            json={"visible_to_members": False},
        )
        assert hide_response.status_code == 200
        assert hide_response.json()["visible_to_members"] is False


def test_llm_request_error_returns_user_safe_api_error(tmp_path: Path) -> None:
    class FailingSummaryLlmClient(FakeLlmClient):
        def generate_document_summary(self, profile, *, model=None, provider=None):
            _ = profile, model, provider
            raise LlmRequestError(
                stage="summary",
                model="private-model",
                provider="private-provider",
                duration_seconds=1.25,
                cause=RuntimeError("provider secret failure"),
            )

    workbook_path = tmp_path / "standards.xlsx"
    _write_xlsx_fixture(workbook_path)
    with _client_with_llm(tmp_path, FailingSummaryLlmClient()) as client:
        with workbook_path.open("rb") as workbook_file:
            upload_response = client.post(
                "/api/excel/files",
                files={
                    "file": (
                        "standards.xlsx",
                        workbook_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        version_id = upload_response.json()["version"]["version_id"]

        response = client.post(f"/api/excel/versions/{version_id}/summary/generate")

    assert response.status_code == 502
    assert response.json() == {
        "code": "LLM_REQUEST_FAILED",
        "detail": (
            "The summary model request failed. Check the selected model or try again shortly."
        ),
        "retryable": True,
    }


def test_document_summary_can_be_updated(
    client: TestClient,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "standards.xlsx"
    _write_xlsx_fixture(workbook_path)
    with workbook_path.open("rb") as workbook_file:
        upload_response = client.post(
            "/api/excel/files",
            files={
                "file": (
                    "standards.xlsx",
                    workbook_file,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert upload_response.status_code == 200
    version_id = upload_response.json()["version"]["version_id"]

    summary_response = client.post(f"/api/excel/versions/{version_id}/summary/generate")
    assert summary_response.status_code == 200

    update_response = client.patch(
        f"/api/excel/versions/{version_id}/summary",
        json={
            "document_title": "Editable generated title",
            "summary_text": "Updated workbook summary.",
            "business_domain": "standards operations",
            "key_topics": ["Standards", "Standards", "DOW", ""],
            "suitable_questions": ["Which standards are listed?"],
            "routing_notes": "",
        },
    )
    assert update_response.status_code == 200
    updated_summary = update_response.json()
    assert updated_summary["document_title"] == "Editable generated title"
    assert updated_summary["summary_text"] == "Updated workbook summary."
    assert updated_summary["business_domain"] == "standards operations"
    assert updated_summary["key_topics"] == ["Standards", "DOW"]
    assert updated_summary["suitable_questions"] == ["Which standards are listed?"]
    assert updated_summary["routing_notes"] == ""

    persisted_response = client.get(f"/api/excel/versions/{version_id}/summary")
    assert persisted_response.status_code == 200
    assert persisted_response.json()["document_title"] == "Editable generated title"
    assert persisted_response.json()["key_topics"] == ["Standards", "DOW"]


def test_chat_session_sends_all_rows_and_deduplicates_attached_file(
    tmp_path: Path,
) -> None:
    llm_client = CapturingLlmClient()
    with _client_with_llm(tmp_path, llm_client) as client:
        workbook_path = tmp_path / "standards.xlsx"
        _write_large_xlsx_fixture(workbook_path, rows=220)
        with workbook_path.open("rb") as workbook_file:
            upload_response = client.post(
                "/api/excel/files",
                files={
                    "file": (
                        "standards.xlsx",
                        workbook_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        assert upload_response.status_code == 200
        version_id = upload_response.json()["version"]["version_id"]
        summary_response = client.post(f"/api/excel/versions/{version_id}/summary/generate")
        assert summary_response.status_code == 200

        session_response = client.post("/api/excel/chat/sessions")
        assert session_response.status_code == 200
        session_id = session_response.json()["session_id"]

        first_response = client.post(
            f"/api/excel/chat/sessions/{session_id}/messages",
            json={"question": "What standards are listed?"},
        )
        assert first_response.status_code == 200
        first_answer = first_response.json()
        assert first_answer["session_id"] == session_id
        assert first_answer["newly_attached_documents"][0]["version_id"] == version_id
        assert first_answer["attached_documents"][0]["row_count"] == 221
        assert first_answer["citations"][0]["evidence_id"].endswith("::S001_R205")
        assert first_answer["citations"][0]["row_id"] == "S001_R205"
        assert len(llm_client.answer_calls[0]["rows"]) == 221

        history_response = client.get(f"/api/excel/chat/sessions/{session_id}/turns")
        assert history_response.status_code == 200
        turns = history_response.json()["turns"]
        assert len(turns) == 1
        assert turns[0]["question"] == "What standards are listed?"
        assert turns[0]["answer"]["answer_blocks"] == first_answer["answer_blocks"]
        assert turns[0]["answer"]["citations"] == first_answer["citations"]
        assert turns[0]["answer"]["selected_documents"] == first_answer["selected_documents"]

        second_response = client.post(
            f"/api/excel/chat/sessions/{session_id}/messages",
            json={"question": "What is the date for that?"},
        )
        assert second_response.status_code == 200
        second_answer = second_response.json()
        assert second_answer["newly_attached_documents"] == []
        assert len(second_answer["attached_documents"]) == 1
        assert llm_client.route_calls[1]["user_questions"] == [
            "What standards are listed?",
            "What is the date for that?",
        ]
        assert len(llm_client.route_calls[1]["attached_documents"]) == 1
        assert len(llm_client.answer_calls[1]["previous_turns"]) == 1

        persisted_history_response = client.get(f"/api/excel/chat/sessions/{session_id}/turns")
        assert persisted_history_response.status_code == 200
        persisted_turns = persisted_history_response.json()["turns"]
        assert [turn["question"] for turn in persisted_turns] == [
            "What standards are listed?",
            "What is the date for that?",
        ]
        assert persisted_turns[0]["answer"]["citations"][0]["row"] == (
            first_answer["citations"][0]["row"]
        )


def test_parallel_questions_for_same_session_are_serialized(
    tmp_path: Path,
) -> None:
    barrier = Barrier(2)
    llm_client = BlockingCapturingLlmClient(barrier)
    with _client_with_llm(tmp_path, llm_client) as client:
        workbook_path = tmp_path / "standards.xlsx"
        _write_large_xlsx_fixture(workbook_path, rows=3)
        with workbook_path.open("rb") as workbook_file:
            upload_response = client.post(
                "/api/excel/files",
                files={
                    "file": (
                        "standards.xlsx",
                        workbook_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        assert upload_response.status_code == 200
        version_id = upload_response.json()["version"]["version_id"]
        assert client.post(f"/api/excel/versions/{version_id}/summary/generate").status_code == 200
        session_id = client.post("/api/excel/chat/sessions").json()["session_id"]

        responses: dict[str, int] = {}

        def ask(question: str) -> None:
            response = client.post(
                f"/api/excel/chat/sessions/{session_id}/messages",
                json={"question": question},
            )
            responses[question] = response.status_code

        first_thread = Thread(target=ask, args=("First concurrent question?",))
        second_thread = Thread(target=ask, args=("Second concurrent question?",))
        first_thread.start()
        barrier.wait(timeout=3)
        second_thread.start()
        first_thread.join(timeout=5)
        second_thread.join(timeout=5)

        assert responses == {
            "First concurrent question?": 200,
            "Second concurrent question?": 200,
        }
        assert [call["question"] for call in llm_client.route_calls] == [
            "First concurrent question?",
            "Second concurrent question?",
        ]
        assert llm_client.route_calls[1]["user_questions"] == [
            "First concurrent question?",
            "Second concurrent question?",
        ]
        turns = client.get(f"/api/excel/chat/sessions/{session_id}/turns").json()["turns"]
        assert [turn["question"] for turn in turns] == [
            "First concurrent question?",
            "Second concurrent question?",
        ]


def test_file_referenced_by_chat_attachment_can_be_deleted_without_breaking_chat(
    tmp_path: Path,
) -> None:
    llm_client = CapturingLlmClient()
    with _client_with_llm(tmp_path, llm_client) as client:
        workbook_path = tmp_path / "standards.xlsx"
        _write_large_xlsx_fixture(workbook_path, rows=5)
        with workbook_path.open("rb") as workbook_file:
            upload_response = client.post(
                "/api/excel/files",
                files={
                    "file": (
                        "standards.xlsx",
                        workbook_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        assert upload_response.status_code == 200
        file_id = upload_response.json()["file"]["file_id"]
        version_id = upload_response.json()["version"]["version_id"]
        assert client.post(f"/api/excel/versions/{version_id}/summary/generate").status_code == 200

        session_id = client.post("/api/excel/chat/sessions").json()["session_id"]
        answer_response = client.post(
            f"/api/excel/chat/sessions/{session_id}/messages",
            json={"question": "Attach this workbook."},
        )
        assert answer_response.status_code == 200

        delete_response = client.delete(f"/api/excel/files/{file_id}?confirm_delete=true")
        assert delete_response.status_code == 200
        assert delete_response.json()["file_id"] == file_id

        file_response = client.get(f"/api/excel/files/{file_id}")
        assert file_response.status_code == 404
        summary_after_delete_response = client.get(f"/api/excel/versions/{version_id}/summary")
        assert summary_after_delete_response.status_code == 404
        list_response = client.get("/api/excel/files")
        assert list_response.status_code == 200
        assert list_response.json()["files"] == []

        follow_up_response = client.post(
            f"/api/excel/chat/sessions/{session_id}/messages",
            json={"question": "Can this session still use the workbook?"},
        )
        assert follow_up_response.status_code == 200
        assert follow_up_response.json()["selected_documents"][0]["version_id"] == version_id
        assert len(llm_client.answer_calls[-1]["rows"]) == 6


def test_chat_answer_row_limit_caps_loaded_rows_and_persists_warning(
    tmp_path: Path,
) -> None:
    llm_client = CapturingLlmClient()
    with _client_with_llm(tmp_path, llm_client, max_answer_rows=3) as client:
        workbook_path = tmp_path / "standards.xlsx"
        _write_large_xlsx_fixture(workbook_path, rows=10)
        with workbook_path.open("rb") as workbook_file:
            upload_response = client.post(
                "/api/excel/files",
                files={
                    "file": (
                        "standards.xlsx",
                        workbook_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        assert upload_response.status_code == 200
        version_id = upload_response.json()["version"]["version_id"]
        summary_response = client.post(f"/api/excel/versions/{version_id}/summary/generate")
        assert summary_response.status_code == 200

        session_response = client.post("/api/excel/chat/sessions")
        assert session_response.status_code == 200
        session_id = session_response.json()["session_id"]

        answer_response = client.post(
            f"/api/excel/chat/sessions/{session_id}/messages",
            json={"question": "What standards are listed?"},
        )

        assert answer_response.status_code == 200
        answer = answer_response.json()
        assert len(llm_client.answer_calls[0]["rows"]) == 3
        assert answer["warnings"] == [
            (
                "Only the first 3 row(s) were inspected to keep the answer request "
                "within the long-running safety limit."
            )
        ]

        history_response = client.get(f"/api/excel/chat/sessions/{session_id}/turns")
        assert history_response.status_code == 200
        assert history_response.json()["turns"][0]["answer"]["warnings"] == answer["warnings"]


def test_chat_session_can_be_listed_renamed_pinned_and_deleted(
    tmp_path: Path,
) -> None:
    llm_client = CapturingLlmClient()
    with _client_with_llm(tmp_path, llm_client) as client:
        first_response = client.post("/api/excel/chat/sessions")
        second_response = client.post("/api/excel/chat/sessions")
        assert first_response.status_code == 200
        assert second_response.status_code == 200
        first_session_id = first_response.json()["session_id"]
        second_session_id = second_response.json()["session_id"]

        rename_response = client.patch(
            f"/api/excel/chat/sessions/{first_session_id}",
            json={"title": "Financial Analysis Q3"},
        )
        assert rename_response.status_code == 200
        renamed = rename_response.json()
        assert renamed["title"] == "Financial Analysis Q3"
        assert renamed["pinned_at"] is None

        pin_response = client.patch(
            f"/api/excel/chat/sessions/{first_session_id}/pin",
            json={"pinned": True},
        )
        assert pin_response.status_code == 200
        pinned = pin_response.json()
        assert pinned["pinned_at"] is not None

        list_response = client.get("/api/excel/chat/sessions")
        assert list_response.status_code == 200
        sessions = list_response.json()["sessions"]
        assert [session["session_id"] for session in sessions] == [
            first_session_id,
            second_session_id,
        ]

        unpin_response = client.patch(
            f"/api/excel/chat/sessions/{first_session_id}/pin",
            json={"pinned": False},
        )
        assert unpin_response.status_code == 200
        assert unpin_response.json()["pinned_at"] is None

        delete_response = client.delete(f"/api/excel/chat/sessions/{first_session_id}")
        assert delete_response.status_code == 204

        missing_response = client.get(f"/api/excel/chat/sessions/{first_session_id}")
        assert missing_response.status_code == 404


def test_chat_route_returns_documents_before_answer_stage(
    tmp_path: Path,
) -> None:
    llm_client = CapturingLlmClient()
    with _client_with_llm(tmp_path, llm_client) as client:
        workbook_path = tmp_path / "standards.xlsx"
        _write_large_xlsx_fixture(workbook_path, rows=5)
        with workbook_path.open("rb") as workbook_file:
            upload_response = client.post(
                "/api/excel/files",
                files={
                    "file": (
                        "standards.xlsx",
                        workbook_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        version_id = upload_response.json()["version"]["version_id"]
        client.post(f"/api/excel/versions/{version_id}/summary/generate")
        session_id = client.post("/api/excel/chat/sessions").json()["session_id"]
        session_before_route = client.get(
            f"/api/excel/chat/sessions/{session_id}"
        ).json()

        route_response = client.post(
            f"/api/excel/chat/sessions/{session_id}/route",
            json={
                "question": "route first",
                "request_id": "route-plan-request",
            },
        )
        assert route_response.status_code == 200
        route_payload = route_response.json()
        assert route_payload["selected_documents"][0]["version_id"] == version_id
        assert route_payload["newly_attached_documents"][0]["version_id"] == version_id
        assert route_payload["attached_documents"][0]["row_count"] == 6
        assert route_payload["request_id"] == "route-plan-request"
        assert route_payload["session_revision"] == 0
        assert "timings" not in route_payload
        assert llm_client.answer_calls == []
        assert client.get(
            f"/api/excel/chat/sessions/{session_id}"
        ).json() == session_before_route
        assert client.get(
            f"/api/excel/chat/sessions/{session_id}/turns"
        ).json()["turns"] == []

        answer_response = client.post(
            f"/api/excel/chat/sessions/{session_id}/answer",
            json={
                "question": "route first",
                "selected_version_ids": [
                    document["version_id"]
                    for document in route_payload["selected_documents"]
                ],
                "session_revision": route_payload["session_revision"],
                "request_id": "route-answer-request",
            },
        )
        assert answer_response.status_code == 200
        answer_payload = answer_response.json()
        assert answer_payload["newly_attached_documents"][0]["version_id"] == version_id
        assert answer_payload["answer_blocks"][0]["citation_ids"] == ["C1"]
        assert "timings" not in answer_payload


def test_chat_answer_request_passes_deep_thinking_flag(
    tmp_path: Path,
) -> None:
    llm_client = CapturingLlmClient()
    with _client_with_llm(tmp_path, llm_client) as client:
        workbook_path = tmp_path / "standards.xlsx"
        _write_large_xlsx_fixture(workbook_path, rows=5)
        with workbook_path.open("rb") as workbook_file:
            upload_response = client.post(
                "/api/excel/files",
                files={
                    "file": (
                        "standards.xlsx",
                        workbook_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        version_id = upload_response.json()["version"]["version_id"]
        client.post(f"/api/excel/versions/{version_id}/summary/generate")
        session_id = client.post("/api/excel/chat/sessions").json()["session_id"]

        response = client.post(
            f"/api/excel/chat/sessions/{session_id}/messages",
            json={
                "question": "Think carefully about the standards.",
                "enable_deep_thinking": True,
            },
        )

        assert response.status_code == 200
        assert response.json()["answer_blocks"][0]["reasoning"] == "Captured reasoning."
        assert llm_client.answer_calls[0]["enable_deep_thinking"] is True


def test_chat_request_id_is_idempotent_and_payload_bound(
    tmp_path: Path,
) -> None:
    llm_client = CapturingLlmClient()
    with _client_with_llm(tmp_path, llm_client) as client:
        session_id = client.post("/api/excel/chat/sessions").json()["session_id"]
        path = f"/api/excel/chat/sessions/{session_id}/messages"
        request_payload = {
            "question": "No documents are available.",
            "request_id": "excel-idempotency-request",
        }

        first_response = client.post(path, json=request_payload)
        second_response = client.post(path, json=request_payload)

        assert first_response.status_code == 200
        assert second_response.status_code == 200
        assert second_response.json() == first_response.json()
        assert len(llm_client.route_calls) == 1
        turns = client.get(
            f"/api/excel/chat/sessions/{session_id}/turns"
        ).json()["turns"]
        assert len(turns) == 1

        conflict_response = client.post(
            path,
            json={
                "question": "A different payload must not reuse the request ID.",
                "request_id": "excel-idempotency-request",
            },
        )
        assert conflict_response.status_code == 409
        assert conflict_response.json()["code"] == "CHAT_IDEMPOTENCY_CONFLICT"
        assert len(llm_client.route_calls) == 1


def test_failed_chat_request_releases_idempotency_claim_for_retry(
    tmp_path: Path,
) -> None:
    class FailOnceRouteLlmClient(CapturingLlmClient):
        def __init__(self) -> None:
            super().__init__()
            self.failed = False

        def route_documents(self, *args, **kwargs) -> list[SelectedDocument]:
            if not self.failed:
                self.failed = True
                raise LlmRequestError(
                    stage="router",
                    model="test-router",
                    provider="test-provider",
                    duration_seconds=0.1,
                    cause=RuntimeError("injected transient failure"),
                )
            return super().route_documents(*args, **kwargs)

    llm_client = FailOnceRouteLlmClient()
    with _client_with_llm(tmp_path, llm_client) as client:
        session_id = client.post("/api/excel/chat/sessions").json()["session_id"]
        path = f"/api/excel/chat/sessions/{session_id}/messages"
        payload = {
            "question": "Retry this request safely.",
            "request_id": "retry-after-failure-request",
        }

        failed_response = client.post(path, json=payload)
        retried_response = client.post(path, json=payload)

        assert failed_response.status_code == 502
        assert retried_response.status_code == 200
        turns = client.get(
            f"/api/excel/chat/sessions/{session_id}/turns"
        ).json()["turns"]
        assert [turn["question"] for turn in turns] == [payload["question"]]


def test_unknown_excel_chat_session_is_never_created_implicitly(
    tmp_path: Path,
) -> None:
    llm_client = CapturingLlmClient()
    with _client_with_llm(tmp_path, llm_client) as client:
        sessions_before = client.get("/api/excel/chat/sessions").json()["sessions"]
        missing_session_id = "session_missing_excel_chat"

        message_response = client.post(
            f"/api/excel/chat/sessions/{missing_session_id}/messages",
            json={
                "question": "Do not create this session.",
                "request_id": "missing-message-request",
            },
        )
        route_response = client.post(
            f"/api/excel/chat/sessions/{missing_session_id}/route",
            json={"question": "Do not create this session."},
        )
        answer_response = client.post(
            f"/api/excel/chat/sessions/{missing_session_id}/answer",
            json={
                "question": "Do not create this session.",
                "request_id": "missing-answer-request",
            },
        )

        assert message_response.status_code == 404
        assert route_response.status_code == 404
        assert answer_response.status_code == 404
        assert client.get("/api/excel/chat/sessions").json()["sessions"] == sessions_before
        assert llm_client.route_calls == []
        assert llm_client.answer_calls == []


def test_stale_route_revision_cannot_commit_excel_chat_turn(
    tmp_path: Path,
) -> None:
    llm_client = CapturingLlmClient()
    with _client_with_llm(tmp_path, llm_client) as client:
        workbook_path = tmp_path / "revision.xlsx"
        _write_large_xlsx_fixture(workbook_path, rows=3)
        with workbook_path.open("rb") as workbook_file:
            upload_response = client.post(
                "/api/excel/files",
                files={
                    "file": (
                        "revision.xlsx",
                        workbook_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        version_id = upload_response.json()["version"]["version_id"]
        client.post(f"/api/excel/versions/{version_id}/summary/generate")
        session_id = client.post("/api/excel/chat/sessions").json()["session_id"]

        route_payload = client.post(
            f"/api/excel/chat/sessions/{session_id}/route",
            json={"question": "Prepare a stale route."},
        ).json()
        committed_response = client.post(
            f"/api/excel/chat/sessions/{session_id}/messages",
            json={
                "question": "Commit a newer turn first.",
                "request_id": "newer-turn-request",
            },
        )
        assert committed_response.status_code == 200

        stale_response = client.post(
            f"/api/excel/chat/sessions/{session_id}/answer",
            json={
                "question": "Prepare a stale route.",
                "selected_version_ids": [version_id],
                "session_revision": route_payload["session_revision"],
                "request_id": "stale-route-answer-request",
            },
        )

        assert stale_response.status_code == 409
        assert stale_response.json()["code"] == "CHAT_SESSION_REVISION_CONFLICT"
        assert len(llm_client.answer_calls) == 1
        turns = client.get(
            f"/api/excel/chat/sessions/{session_id}/turns"
        ).json()["turns"]
        assert [turn["question"] for turn in turns] == ["Commit a newer turn first."]


def test_langgraph_workflow_runs_full_chat_chain(
    tmp_path: Path,
) -> None:
    llm_client = CapturingLlmClient()
    with _client_with_llm(
        tmp_path,
        llm_client,
        workflow=LangGraphChatWorkflow(),
    ) as client:
        workbook_path = tmp_path / "standards.xlsx"
        _write_large_xlsx_fixture(workbook_path, rows=5)
        with workbook_path.open("rb") as workbook_file:
            upload_response = client.post(
                "/api/excel/files",
                files={
                    "file": (
                        "standards.xlsx",
                        workbook_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        version_id = upload_response.json()["version"]["version_id"]
        client.post(f"/api/excel/versions/{version_id}/summary/generate")

        chat_response = client.post(
            "/api/excel/chat",
            json={"question": "What standards are listed?"},
        )

        assert chat_response.status_code == 200
        answer_payload = chat_response.json()
        assert llm_client.route_calls[0]["question"] == "What standards are listed?"
        assert llm_client.answer_calls[0]["question"] == "What standards are listed?"
        assert answer_payload["selected_documents"][0]["version_id"] == version_id
        assert "timings" not in answer_payload


def test_llm_options_endpoint_and_workspace_models(
    tmp_path: Path,
) -> None:
    llm_client = CapturingLlmClient()
    with _client_with_llm(tmp_path, llm_client) as client:
        options_response = client.get("/api/excel/llm/options")
        assert options_response.status_code == 200
        options_payload = options_response.json()
        assert "deepseek-ai/DeepSeek-V4-Pro" in options_payload["models"]
        assert "Qwen/Qwen3.6-27B" in options_payload["models"]
        assert options_payload["models"][0] == "inclusionAI/Ling-flash-2.0"
        providers_by_id = {
            provider["provider"]: provider
            for provider in options_payload["providers"]
        }
        assert providers_by_id["deepseek"]["deep_thinking_models"] == [
            "deepseek-v4-pro",
            "deepseek-v4-flash",
        ]
        assert providers_by_id["volcengine_ark"]["models"][:3] == [
            "doubao-seed-2-0-pro-260215",
            "doubao-seed-2-0-lite-260428",
            "doubao-seed-2-0-mini-260428",
        ]
        assert "deepseek-v4-pro-260425" in providers_by_id["volcengine_ark"]["models"]
        assert providers_by_id["volcengine_ark"]["deep_thinking_models"] == []
        assert providers_by_id["siliconflow"]["deep_thinking_models"] == [
            "Pro/deepseek-ai/DeepSeek-V3.2",
        ]
        assert options_payload["defaults"] == {
            "summary_provider": "deepseek",
            "summary_model": "deepseek-v4-pro",
            "router_provider": "siliconflow",
            "router_model": "Qwen/Qwen3.6-35B-A3B",
            "answer_provider": "deepseek",
            "answer_model": "deepseek-v4-pro",
        }

        workbook_path = tmp_path / "standards.xlsx"
        _write_large_xlsx_fixture(workbook_path, rows=5)
        with workbook_path.open("rb") as workbook_file:
            upload_response = client.post(
                "/api/excel/files",
                files={
                    "file": (
                        "standards.xlsx",
                        workbook_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        version_id = upload_response.json()["version"]["version_id"]

        save_response = client.patch(
            "/api/excel/llm/preferences",
            json={
                "summary_provider": "siliconflow",
                "summary_model": "Qwen/Qwen3.6-27B",
                "router_provider": "deepseek",
                "router_model": "deepseek-v4-flash",
                "answer_provider": "siliconflow",
                "answer_model": "deepseek-ai/DeepSeek-V4-Pro",
            },
        )
        assert save_response.status_code == 200

        updated_options_response = client.get("/api/excel/llm/options")
        assert updated_options_response.status_code == 200
        assert updated_options_response.json()["defaults"]["router_model"] == "deepseek-v4-flash"

        summary_response = client.post(f"/api/excel/versions/{version_id}/summary/generate")
        assert summary_response.status_code == 200

        session_id = client.post("/api/excel/chat/sessions").json()["session_id"]
        route_response = client.post(
            f"/api/excel/chat/sessions/{session_id}/route",
            json={
                "question": "route first",
                "router_model": "inclusionAI/Ling-flash-2.0",
            },
        )
        assert route_response.status_code == 200
        selected_version_ids = [
            document["version_id"] for document in route_response.json()["selected_documents"]
        ]

        answer_response = client.post(
            f"/api/excel/chat/sessions/{session_id}/answer",
            json={
                "question": "route first",
                "answer_model": "deepseek-ai/DeepSeek-V4-Pro",
                "selected_version_ids": selected_version_ids,
            },
        )
        assert answer_response.status_code == 200

        assert llm_client.summary_models == ["Qwen/Qwen3.6-27B"]
        assert llm_client.route_models == ["deepseek-v4-flash"]
        assert llm_client.answer_models == ["deepseek-ai/DeepSeek-V4-Pro"]


def test_chat_turn_uses_model_preference_snapshot_when_defaults_change_mid_turn(
    tmp_path: Path,
) -> None:
    repository = SQLiteExcelAssetRepository(tmp_path / "excel.sqlite3")
    excel_assets = ExcelAssetService(
        repository=repository,
        storage=FilesystemExcelArtifactStorage(tmp_path / "storage"),
        workbook_reader=OpenpyxlWorkbookReader(),
    )
    excel_assets.initialize()
    llm_preferences = WorkspaceLlmPreferenceService(
        repository=repository,
        settings=Settings(_env_file=None),
    )
    llm_preferences.save_preference(
        summary_provider="initial-summary-provider",
        summary_model="initial-summary-model",
        router_provider="initial-router-provider",
        router_model="initial-router-model",
        answer_provider="initial-answer-provider",
        answer_model="initial-answer-model",
    )
    llm_client = PreferenceMutatingLlmClient(llm_preferences)
    summaries = DocumentSummaryService(
        excel_assets=excel_assets,
        llm_client=llm_client,
        repository=repository,
        llm_preferences=llm_preferences,
    )
    chat = ChatService(
        excel_assets=excel_assets,
        summaries=summaries,
        llm_client=llm_client,
        sessions=repository,
        llm_preferences=llm_preferences,
        workflow=LangGraphChatWorkflow(),
    )
    workbook_path = tmp_path / "standards.xlsx"
    _write_large_xlsx_fixture(workbook_path, rows=5)
    upload = excel_assets.upload_workbook("standards.xlsx", workbook_path.read_bytes())
    summaries.generate_summary(upload.version.version_id)

    answer = chat.answer_question("Which standards are listed?")

    assert answer.selected_documents[0].version_id == upload.version.version_id
    assert llm_client.route_models[-1] == "initial-router-model"
    assert llm_client.answer_models[-1] == "initial-answer-model"
    assert llm_preferences.get_preference().answer_model == "next-answer-model"


def test_cancelled_chat_turn_is_not_persisted_or_used_as_context(
    tmp_path: Path,
) -> None:
    repository = SQLiteExcelAssetRepository(tmp_path / "excel.sqlite3")
    excel_assets = ExcelAssetService(
        repository=repository,
        storage=FilesystemExcelArtifactStorage(tmp_path / "storage"),
        workbook_reader=OpenpyxlWorkbookReader(),
    )
    excel_assets.initialize()
    llm_preferences = WorkspaceLlmPreferenceService(
        repository=repository,
        settings=Settings(_env_file=None),
    )
    cancellation = ChatCancellationToken(request_id="chat-cancel-test")
    llm_client = CancellingAfterRouteLlmClient(cancellation)
    summaries = DocumentSummaryService(
        excel_assets=excel_assets,
        llm_client=llm_client,
        repository=repository,
        llm_preferences=llm_preferences,
    )
    chat = ChatService(
        excel_assets=excel_assets,
        summaries=summaries,
        llm_client=llm_client,
        sessions=repository,
        llm_preferences=llm_preferences,
    )
    workbook_path = tmp_path / "standards.xlsx"
    _write_large_xlsx_fixture(workbook_path, rows=5)
    upload = excel_assets.upload_workbook("standards.xlsx", workbook_path.read_bytes())
    summaries.generate_summary(upload.version.version_id)
    session = chat.create_session_for_user("user_1")

    with pytest.raises(ChatRequestCancelledError):
        chat.answer_question(
            "This turn will be cancelled.",
            session_id=session.session_id,
            user_id="user_1",
            cancellation_token=cancellation,
        )

    assert chat.list_turns(session.session_id, user_id="user_1") == []
    assert repository.list_attached_documents(session.session_id) == []

    llm_client.cancel_after_route = False
    cancellation = ChatCancellationToken(request_id="chat-next-turn")
    answer = chat.answer_question(
        "Next turn should not see cancelled context.",
        session_id=session.session_id,
        user_id="user_1",
        cancellation_token=cancellation,
    )

    assert answer.question == "Next turn should not see cancelled context."
    assert llm_client.answer_calls[-1]["previous_turns"] == []


def test_chat_cancellation_is_shared_across_registries(tmp_path: Path) -> None:
    repository = SQLiteExcelAssetRepository(tmp_path / "excel.sqlite3")
    repository.initialize()
    cancelling_registry = ChatCancellationRegistry(repository=repository)
    answering_registry = ChatCancellationRegistry(repository=repository)

    cancelled_immediately = cancelling_registry.cancel("chat-shared-cancel")
    token = answering_registry.register("chat-shared-cancel")

    assert cancelled_immediately is True
    assert token is not None
    with pytest.raises(ChatRequestCancelledError):
        token.raise_if_cancelled()


def test_hidden_file_is_not_routed_or_used_for_member_chat(
    tmp_path: Path,
) -> None:
    repository = SQLiteExcelAssetRepository(tmp_path / "excel.sqlite3")
    excel_assets = ExcelAssetService(
        repository=repository,
        storage=FilesystemExcelArtifactStorage(tmp_path / "storage"),
        workbook_reader=OpenpyxlWorkbookReader(),
    )
    excel_assets.initialize()
    llm_preferences = WorkspaceLlmPreferenceService(
        repository=repository,
        settings=Settings(_env_file=None),
    )
    llm_client = CapturingLlmClient()
    summaries = DocumentSummaryService(
        excel_assets=excel_assets,
        llm_client=llm_client,
        repository=repository,
        llm_preferences=llm_preferences,
    )
    chat = ChatService(
        excel_assets=excel_assets,
        summaries=summaries,
        llm_client=llm_client,
        sessions=repository,
        llm_preferences=llm_preferences,
    )

    public_path = tmp_path / "public.xlsx"
    hidden_path = tmp_path / "hidden.xlsx"
    _write_xlsx_fixture(public_path)
    _write_xlsx_fixture(hidden_path)
    public_upload = excel_assets.upload_workbook("public.xlsx", public_path.read_bytes())
    hidden_upload = excel_assets.upload_workbook("hidden.xlsx", hidden_path.read_bytes())
    summaries.generate_summary(public_upload.version.version_id)
    summaries.generate_summary(hidden_upload.version.version_id)
    excel_assets.set_file_visibility(hidden_upload.file.file_id, visible_to_members=False)

    answer = chat.answer_question(
        "Which standards are available?",
        user_id="user_member_test",
        user_role=member_user().role,
    )

    routed_summary_versions = [
        summary.version_id
        for call in llm_client.route_calls
        for summary in call["summaries"]
    ]
    assert hidden_upload.version.version_id not in routed_summary_versions
    assert answer.selected_documents[0].version_id == public_upload.version.version_id


def test_member_chat_revalidates_visibility_after_model_answer(
    tmp_path: Path,
) -> None:
    repository = SQLiteExcelAssetRepository(tmp_path / "excel.sqlite3")
    excel_assets = ExcelAssetService(
        repository=repository,
        storage=FilesystemExcelArtifactStorage(tmp_path / "storage"),
        workbook_reader=OpenpyxlWorkbookReader(),
    )
    excel_assets.initialize()
    llm_preferences = WorkspaceLlmPreferenceService(
        repository=repository,
        settings=Settings(_env_file=None),
    )
    llm_client = VisibilityHidingLlmClient(excel_assets)
    summaries = DocumentSummaryService(
        excel_assets=excel_assets,
        llm_client=llm_client,
        repository=repository,
        llm_preferences=llm_preferences,
    )
    chat = ChatService(
        excel_assets=excel_assets,
        summaries=summaries,
        llm_client=llm_client,
        sessions=repository,
        llm_preferences=llm_preferences,
    )

    workbook_path = tmp_path / "temporary-visible.xlsx"
    _write_xlsx_fixture(workbook_path)
    upload = excel_assets.upload_workbook(
        "temporary-visible.xlsx",
        workbook_path.read_bytes(),
    )
    summaries.generate_summary(upload.version.version_id)
    llm_client.hide_file_id_on_answer = upload.file.file_id

    answer = chat.answer_question(
        "Use the temporary file.",
        user_id="user_member_visibility_race",
        user_role=member_user().role,
    )

    assert len(llm_client.answer_calls) == 1
    assert answer.insufficient_evidence is True
    assert answer.selected_documents == []
    assert answer.citations == []
    assert answer.answer_blocks[0].citation_ids == []


def test_llm_preferences_are_persisted(
    tmp_path: Path,
) -> None:
    llm_client = CapturingLlmClient()
    with _client_with_llm(tmp_path, llm_client) as client:
        default_response = client.get("/api/excel/llm/preferences")
        assert default_response.status_code == 200
        assert default_response.json()["summary_provider"] == "deepseek"

        save_response = client.patch(
            "/api/excel/llm/preferences",
            json={
                "summary_provider": "siliconflow",
                "summary_model": "Qwen/Qwen3.6-27B",
                "router_provider": "deepseek",
                "router_model": "deepseek-v4-flash",
                "answer_provider": "siliconflow",
                "answer_model": "deepseek-ai/DeepSeek-V4-Pro",
            },
        )
        assert save_response.status_code == 200
        saved = save_response.json()
        assert saved["scope"] == "workspace"
        assert saved["summary_model"] == "Qwen/Qwen3.6-27B"
        assert saved["updated_at"]

        persisted_response = client.get("/api/excel/llm/preferences")
        assert persisted_response.status_code == 200
        assert persisted_response.json() == saved


def test_verifier_uses_evidence_id_to_keep_correct_file() -> None:
    chat = ChatService(
        excel_assets=None,  # type: ignore[arg-type]
        summaries=None,  # type: ignore[arg-type]
        llm_client=FakeLlmClient(),
        sessions=None,  # type: ignore[arg-type]
        llm_preferences=_fake_llm_preferences(),
    )
    citation_index = {
        "version_a::sheet_a::S001_R5": ExcelCitation(
            citation_id="",
            evidence_id="version_a::sheet_a::S001_R5",
            file_id="file_a",
            version_id="version_a",
            sheet_id="sheet_a",
            sheet_name="Sheet A",
            row_id="S001_R5",
            row=["S001_R5", "A row"],
        ),
        "version_b::sheet_b::S001_R5": ExcelCitation(
            citation_id="",
            evidence_id="version_b::sheet_b::S001_R5",
            file_id="file_b",
            version_id="version_b",
            sheet_id="sheet_b",
            sheet_name="Sheet B",
            row_id="S001_R5",
            row=["S001_R5", "B row"],
        ),
    }

    citations, evidence_id_to_citation_id, warnings = chat._build_verified_citations(
        [
            DraftCitation(
                evidence_id="version_a::sheet_a::S001_R5",
                quote="A row",
            )
        ],
        ["version_a::sheet_a::S001_R5"],
        citation_index,
    )

    assert warnings == []
    assert evidence_id_to_citation_id == {"version_a::sheet_a::S001_R5": "C1"}
    assert citations[0].file_id == "file_a"
    assert citations[0].sheet_id == "sheet_a"
    assert citations[0].row_id == "S001_R5"


def test_verifier_rejects_ambiguous_legacy_row_id() -> None:
    chat = ChatService(
        excel_assets=None,  # type: ignore[arg-type]
        summaries=None,  # type: ignore[arg-type]
        llm_client=FakeLlmClient(),
        sessions=None,  # type: ignore[arg-type]
        llm_preferences=_fake_llm_preferences(),
    )
    citation_index = {
        "version_a::sheet_a::S001_R5": ExcelCitation(
            citation_id="",
            evidence_id="version_a::sheet_a::S001_R5",
            file_id="file_a",
            version_id="version_a",
            sheet_id="sheet_a",
            sheet_name="Sheet A",
            row_id="S001_R5",
            row=["S001_R5", "A row"],
        ),
        "version_b::sheet_b::S001_R5": ExcelCitation(
            citation_id="",
            evidence_id="version_b::sheet_b::S001_R5",
            file_id="file_b",
            version_id="version_b",
            sheet_id="sheet_b",
            sheet_name="Sheet B",
            row_id="S001_R5",
            row=["S001_R5", "B row"],
        ),
    }

    citations, evidence_id_to_citation_id, warnings = chat._build_verified_citations(
        [DraftCitation(row_id="S001_R5", quote="legacy row id only")],
        ["S001_R5"],
        citation_index,
    )

    assert citations == []
    assert evidence_id_to_citation_id == {}
    assert any("ignored ambiguous citation row_id: S001_R5" == warning for warning in warnings)


def test_verifier_rejects_invalid_evidence_id() -> None:
    chat = ChatService(
        excel_assets=None,  # type: ignore[arg-type]
        summaries=None,  # type: ignore[arg-type]
        llm_client=FakeLlmClient(),
        sessions=None,  # type: ignore[arg-type]
        llm_preferences=_fake_llm_preferences(),
    )
    citation_index = {
        "version_a::sheet_a::S001_R5": ExcelCitation(
            citation_id="",
            evidence_id="version_a::sheet_a::S001_R5",
            file_id="file_a",
            version_id="version_a",
            sheet_id="sheet_a",
            sheet_name="Sheet A",
            row_id="S001_R5",
            row=["S001_R5", "A row"],
        )
    }

    citations, evidence_id_to_citation_id, warnings = chat._build_verified_citations(
        [DraftCitation(evidence_id="version_x::sheet_y::S001_R5", quote="bad")],
        ["version_x::sheet_y::S001_R5"],
        citation_index,
    )

    assert citations == []
    assert evidence_id_to_citation_id == {}
    assert any(
        warning == "ignored invalid citation evidence_id: version_x::sheet_y::S001_R5"
        for warning in warnings
    )


def _write_xlsx_fixture(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Standards"
    worksheet.append(["Code", "Date"])
    worksheet.append(["EN 60335-1:2023", "2024-01-01"])
    workbook.save(path)


def _write_workflow_xlsx_fixture(path: Path) -> None:
    workbook = Workbook()
    suppliers = workbook.active
    suppliers.title = "Suppliers"
    suppliers.append(["Supplier", "Risk", "Region"])
    suppliers.append(["Apex", "High", "North"])
    suppliers.append(["Beacon", "Low", "South"])
    orders = workbook.create_sheet("Orders")
    orders.append(["Order", "Owner"])
    orders.append(["PO-100", "Liu"])
    workbook.save(path)


def _write_large_xlsx_fixture(path: Path, rows: int) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Standards"
    worksheet.append(["Code", "Date"])
    for row_index in range(1, rows + 1):
        worksheet.append([f"EN 60335-{row_index}", f"2024-01-{row_index % 28 + 1:02d}"])
    workbook.save(path)


class CapturingLlmClient(FakeLlmClient):
    def __init__(self) -> None:
        self.summary_models: list[str | None] = []
        self.route_models: list[str | None] = []
        self.answer_models: list[str | None] = []
        self.route_calls: list[dict] = []
        self.answer_calls: list[dict] = []

    def generate_document_summary(
        self,
        profile,
        *,
        model: str | None = None,
        provider: str | None = None,
    ):
        _ = provider
        self.summary_models.append(model)
        return super().generate_document_summary(profile, model=model, provider=provider)

    def route_documents(
        self,
        question: str,
        summaries: list[DocumentSummary],
        max_documents: int,
        user_questions: list[str] | None = None,
        attached_documents: list[AttachedDocument] | None = None,
        previous_turns: list[ChatTurn] | None = None,
        model: str | None = None,
        provider: str | None = None,
        cancellation_checker=None,
    ) -> list[SelectedDocument]:
        _ = provider, cancellation_checker
        self.route_models.append(model)
        self.route_calls.append(
            {
                "question": question,
                "user_questions": user_questions or [],
                "attached_documents": attached_documents or [],
                "previous_turns": previous_turns or [],
                "summaries": summaries,
            }
        )
        if not summaries:
            return []
        summary = summaries[0]
        return [
            SelectedDocument(
                file_id=summary.file_id,
                version_id=summary.version_id,
                reason="captured test selection",
                confidence=1.0,
            )
        ]

    def answer_with_rows(
        self,
        question: str,
        documents: list[SelectedDocument],
        rows: list[dict],
        previous_turns: list[ChatTurn] | None = None,
        model: str | None = None,
        provider: str | None = None,
        enable_deep_thinking: bool = False,
        cancellation_checker=None,
    ) -> DraftChatAnswer:
        _ = provider, cancellation_checker
        self.answer_models.append(model)
        self.answer_calls.append(
            {
                "question": question,
                "documents": documents,
                "rows": rows,
                "previous_turns": previous_turns or [],
                "enable_deep_thinking": enable_deep_thinking,
            }
        )
        available_row_ids = {str(row["row_id"]) for row in rows}
        row_id = (
            "S001_R205"
            if "S001_R205" in available_row_ids
            else str(rows[-1]["row_id"])
            if rows
            else ""
        )
        evidence_id = ""
        if row_id:
            evidence_id = next(
                str(row["evidence_id"])
                for row in rows
                if str(row["row_id"]) == row_id
            )
        evidence_ids = [evidence_id] if evidence_id else []
        citations = (
            [DraftCitation(evidence_id=evidence_id, quote="captured row")]
            if evidence_id
            else []
        )
        return DraftChatAnswer(
            answer_blocks=[
                DraftAnswerBlock(
                    text=f"Captured answer for {question}.",
                    evidence_ids=evidence_ids,
                    reasoning="Captured reasoning." if enable_deep_thinking else "",
                )
            ],
            citations=citations,
            insufficient_evidence=not rows,
            follow_up_suggestions=[],
        )


class BlockingCapturingLlmClient(CapturingLlmClient):
    def __init__(self, first_route_started: Barrier) -> None:
        super().__init__()
        self._first_route_started = first_route_started
        self._route_count = 0

    def route_documents(self, *args, **kwargs) -> list[SelectedDocument]:
        self._route_count += 1
        if self._route_count == 1:
            self._first_route_started.wait(timeout=3)
        return super().route_documents(*args, **kwargs)


class PreferenceMutatingLlmClient(CapturingLlmClient):
    def __init__(self, preferences: WorkspaceLlmPreferenceService) -> None:
        super().__init__()
        self._preferences = preferences

    def route_documents(
        self,
        question: str,
        summaries: list[DocumentSummary],
        max_documents: int,
        user_questions: list[str] | None = None,
        attached_documents: list[AttachedDocument] | None = None,
        previous_turns: list[ChatTurn] | None = None,
        model: str | None = None,
        provider: str | None = None,
        cancellation_checker=None,
    ) -> list[SelectedDocument]:
        selected_documents = super().route_documents(
            question=question,
            summaries=summaries,
            max_documents=max_documents,
            user_questions=user_questions,
            attached_documents=attached_documents,
            previous_turns=previous_turns,
            model=model,
            provider=provider,
            cancellation_checker=cancellation_checker,
        )
        self._preferences.save_preference(
            summary_provider="next-summary-provider",
            summary_model="next-summary-model",
            router_provider="next-router-provider",
            router_model="next-router-model",
            answer_provider="next-answer-provider",
            answer_model="next-answer-model",
        )
        return selected_documents


class VisibilityHidingLlmClient(CapturingLlmClient):
    def __init__(self, excel_assets: ExcelAssetService) -> None:
        super().__init__()
        self._excel_assets = excel_assets
        self.hide_file_id_on_answer: str | None = None

    def answer_with_rows(self, *args, **kwargs) -> DraftChatAnswer:
        answer = super().answer_with_rows(*args, **kwargs)
        if self.hide_file_id_on_answer is not None:
            self._excel_assets.set_file_visibility(
                self.hide_file_id_on_answer,
                visible_to_members=False,
            )
            self.hide_file_id_on_answer = None
        return answer


class CancellingAfterRouteLlmClient(CapturingLlmClient):
    def __init__(self, cancellation: ChatCancellationToken) -> None:
        super().__init__()
        self._cancellation = cancellation
        self.cancel_after_route = True

    def route_documents(self, *args, **kwargs) -> list[SelectedDocument]:
        selected_documents = super().route_documents(*args, **kwargs)
        if self.cancel_after_route:
            self._cancellation.cancel()
        return selected_documents


@contextmanager
def _client_with_llm(
    tmp_path: Path,
    llm_client: FakeLlmClient,
    workflow: ChatWorkflow | None = None,
    max_answer_rows: int = 20_000,
) -> Iterator[TestClient]:
    repository = SQLiteExcelAssetRepository(tmp_path / "excel.sqlite3")
    excel_assets = ExcelAssetService(
        repository=repository,
        storage=FilesystemExcelArtifactStorage(tmp_path / "storage"),
        workbook_reader=OpenpyxlWorkbookReader(),
    )
    excel_assets.initialize()
    llm_preferences = WorkspaceLlmPreferenceService(
        repository=repository,
        settings=Settings(_env_file=None),
    )
    summaries = DocumentSummaryService(
        excel_assets=excel_assets,
        llm_client=llm_client,
        repository=repository,
        llm_preferences=llm_preferences,
    )
    chat = ChatService(
        excel_assets=excel_assets,
        summaries=summaries,
        llm_client=llm_client,
        sessions=repository,
        llm_preferences=llm_preferences,
        max_answer_rows=max_answer_rows,
        workflow=workflow,
    )
    app.dependency_overrides[get_excel_asset_service] = lambda: excel_assets
    app.dependency_overrides[get_document_summary_service] = lambda: summaries
    app.dependency_overrides[get_chat_service] = lambda: chat
    app.dependency_overrides[get_llm_preference_service] = lambda: llm_preferences
    app.dependency_overrides[get_current_user] = admin_user
    app.dependency_overrides[require_admin_user] = admin_user
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


class _StaticLlmPreferences:
    def get_preference(self) -> LlmPreference:
        settings = Settings(_env_file=None)
        return LlmPreference(
            scope="workspace",
            summary_provider=settings.llm_summary_provider,
            summary_model=settings.llm_summary_model,
            router_provider=settings.llm_router_provider,
            router_model=settings.llm_router_model,
            answer_provider=settings.llm_answer_provider,
            answer_model=settings.llm_answer_model,
            created_at="",
            updated_at="",
        )


def _fake_llm_preferences() -> _StaticLlmPreferences:
    return _StaticLlmPreferences()
