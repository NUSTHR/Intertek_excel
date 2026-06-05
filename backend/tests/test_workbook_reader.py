import builtins
import sys
from datetime import date, datetime
from pathlib import Path
from types import SimpleNamespace

import pytest
import xlwt
from openpyxl import Workbook

from app.adapters.workbook.openpyxl_reader import OpenpyxlWorkbookReader
from app.core.errors import InvalidExcelFileError


class FakeXlsCell:
    def __init__(self, ctype: int, value: object) -> None:
        self.ctype = ctype
        self.value = value


class FakeXlsSheet:
    def __init__(self, rows: list[list[FakeXlsCell]]) -> None:
        self._rows = rows
        self.nrows = len(rows)
        self.ncols = max((len(row) for row in rows), default=0)

    def cell(self, row_index: int, column_index: int) -> FakeXlsCell:
        row = self._rows[row_index]
        if column_index >= len(row):
            return FakeXlsCell(0, "")
        return row[column_index]


class FakeXlsWorkbook:
    datemode = 0

    def __init__(self) -> None:
        self._sheets = {
            "Legacy": FakeXlsSheet(
                [
                    [
                        FakeXlsCell(1, "Code"),
                        FakeXlsCell(1, "DOP"),
                        FakeXlsCell(1, ""),
                    ],
                    [
                        FakeXlsCell(1, "EN 60335-1:2023"),
                        FakeXlsCell(3, 45292),
                        FakeXlsCell(2, 10.0),
                    ],
                    [
                        FakeXlsCell(1, "Flag"),
                        FakeXlsCell(4, 1),
                        FakeXlsCell(5, 23),
                    ],
                    [FakeXlsCell(0, ""), FakeXlsCell(0, ""), FakeXlsCell(0, "")],
                ]
            )
        }
        self.released = False

    def sheet_names(self) -> list[str]:
        return list(self._sheets)

    def sheet_by_name(self, sheet_name: str) -> FakeXlsSheet:
        return self._sheets[sheet_name]

    def release_resources(self) -> None:
        self.released = True


def test_read_real_xlsx_workbook_normalizes_rows_and_values(tmp_path: Path) -> None:
    path = tmp_path / "sample.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Standards"
    worksheet.append(["Code", "日期", "Score"])
    worksheet.append(["EN 60335-1:2023", date(2024, 1, 1), 10.0])
    worksheet.append(["备注", "中文内容", None])
    worksheet.append([None, None, None])
    workbook.save(path)

    data = OpenpyxlWorkbookReader().read(path)

    assert len(data.sheets) == 1
    assert data.sheets[0].sheet_name == "Standards"
    assert data.sheets[0].rows == [
        ["Code", "日期", "Score"],
        ["EN 60335-1:2023", "2024-01-01", "10"],
        ["备注", "中文内容"],
    ]


def test_read_real_xls_workbook_normalizes_rows_and_values(tmp_path: Path) -> None:
    path = tmp_path / "sample.xls"
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet("Legacy")
    date_style = xlwt.easyxf(num_format_str="YYYY-MM-DD")
    rows = [
        ["Code", "日期", "Score"],
        ["EN 60335-1:2023", datetime(2024, 1, 1), 10.0],
        ["备注", "中文内容", ""],
    ]
    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            if isinstance(value, datetime):
                worksheet.write(row_index, column_index, value, date_style)
            else:
                worksheet.write(row_index, column_index, value)
    workbook.save(str(path))

    data = OpenpyxlWorkbookReader().read(path)

    assert len(data.sheets) == 1
    assert data.sheets[0].sheet_name == "Legacy"
    assert data.sheets[0].rows == [
        ["Code", "日期", "Score"],
        ["EN 60335-1:2023", "2024-01-01", "10"],
        ["备注", "中文内容"],
    ]


def test_read_legacy_xls_normalizes_rows_and_values(monkeypatch: pytest.MonkeyPatch) -> None:
    workbook = FakeXlsWorkbook()
    fake_xlrd = SimpleNamespace(
        XL_CELL_EMPTY=0,
        XL_CELL_DATE=3,
        XL_CELL_BOOLEAN=4,
        XL_CELL_ERROR=5,
        error_text_from_code={23: "#REF!"},
        open_workbook=lambda _path, on_demand: workbook,
        xldate=SimpleNamespace(
            xldate_as_datetime=lambda _value, _datemode: datetime(2024, 1, 1)
        ),
    )
    monkeypatch.setitem(sys.modules, "xlrd", fake_xlrd)

    data = OpenpyxlWorkbookReader().read(Path("sample.xls"))

    assert workbook.released is True
    assert len(data.sheets) == 1
    assert data.sheets[0].sheet_name == "Legacy"
    assert data.sheets[0].rows == [
        ["Code", "DOP"],
        ["EN 60335-1:2023", "2024-01-01", "10"],
        ["Flag", "TRUE", "#REF!"],
    ]


def test_read_legacy_xls_reports_missing_dependency(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    original_import = builtins.__import__

    def fail_xlrd_import(name: str, *args: object, **kwargs: object) -> object:
        if name == "xlrd":
            raise ImportError("xlrd missing")
        return original_import(name, *args, **kwargs)

    monkeypatch.delitem(sys.modules, "xlrd", raising=False)
    monkeypatch.setattr(builtins, "__import__", fail_xlrd_import)

    with pytest.raises(InvalidExcelFileError, match="legacy .xls workbook support"):
        OpenpyxlWorkbookReader().read(Path("sample.xls"))
