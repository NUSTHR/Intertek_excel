from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.errors import InvalidExcelFileError
from app.ports.workbook_reader import WorkbookData, WorkbookSheet


class OpenpyxlWorkbookReader:
    def read(self, path: Path) -> WorkbookData:
        if path.suffix.lower() == ".xls":
            return self._read_legacy_xls(path)
        return self._read_open_xml_workbook(path)

    def _read_open_xml_workbook(self, path: Path) -> WorkbookData:
        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
        except InvalidFileException as exc:
            raise InvalidExcelFileError(
                "the uploaded file is not a supported Excel workbook"
            ) from exc
        except Exception as exc:
            raise InvalidExcelFileError("failed to read Excel workbook") from exc

        try:
            sheets: list[WorkbookSheet] = []
            for index, worksheet in enumerate(workbook.worksheets, start=1):
                rows = [
                    self._trim_trailing_empty_cells(
                        [self._to_display_text(cell.value) for cell in row]
                    )
                    for row in worksheet.iter_rows()
                ]
                sheets.append(
                    WorkbookSheet(
                        sheet_index=index,
                        sheet_name=str(worksheet.title or f"Sheet {index}"),
                        rows=self._trim_trailing_empty_rows(rows),
                    )
                )
            return WorkbookData(sheets=sheets)
        finally:
            workbook.close()

    def _read_legacy_xls(self, path: Path) -> WorkbookData:
        try:
            import xlrd
        except ImportError as exc:
            raise InvalidExcelFileError(
                "legacy .xls workbook support is not installed"
            ) from exc

        try:
            workbook = xlrd.open_workbook(str(path), on_demand=True)
        except Exception as exc:
            raise InvalidExcelFileError("failed to read Excel workbook") from exc

        try:
            sheets: list[WorkbookSheet] = []
            for index, sheet_name in enumerate(workbook.sheet_names(), start=1):
                worksheet = workbook.sheet_by_name(sheet_name)
                rows = [
                    self._trim_trailing_empty_cells(
                        [
                            self._to_xls_display_text(
                                worksheet.cell(row_index, column_index),
                                workbook.datemode,
                                xlrd,
                            )
                            for column_index in range(worksheet.ncols)
                        ]
                    )
                    for row_index in range(worksheet.nrows)
                ]
                sheets.append(
                    WorkbookSheet(
                        sheet_index=index,
                        sheet_name=str(sheet_name or f"Sheet {index}"),
                        rows=self._trim_trailing_empty_rows(rows),
                    )
                )
            return WorkbookData(sheets=sheets)
        finally:
            workbook.release_resources()

    def _to_xls_display_text(self, cell: Any, datemode: int, xlrd: Any) -> str:
        if cell.ctype == xlrd.XL_CELL_EMPTY:
            return ""
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return self._to_display_text(
                    xlrd.xldate.xldate_as_datetime(cell.value, datemode)
                )
            except Exception:
                return self._to_display_text(cell.value)
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return "TRUE" if cell.value else "FALSE"
        if cell.ctype == xlrd.XL_CELL_ERROR:
            return xlrd.error_text_from_code.get(cell.value, f"#ERR{cell.value}")
        return self._to_display_text(cell.value)

    def _to_display_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            if value.time() == time():
                return value.date().isoformat()
            return value.isoformat(sep=" ", timespec="seconds")
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, time):
            return value.isoformat(timespec="seconds")
        if isinstance(value, float):
            return self._format_float(value)
        if isinstance(value, Decimal):
            return format(value, "f")
        return str(value)

    def _format_float(self, value: float) -> str:
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")

    def _trim_trailing_empty_cells(self, row: list[str]) -> list[str]:
        last_non_empty_index = -1
        for index, cell in enumerate(row):
            if cell.strip():
                last_non_empty_index = index
        if last_non_empty_index < 0:
            return []
        return row[: last_non_empty_index + 1]

    def _trim_trailing_empty_rows(self, rows: list[list[str]]) -> list[list[str]]:
        last_non_empty_index = -1
        for index, row in enumerate(rows):
            if any(cell.strip() for cell in row):
                last_non_empty_index = index
        if last_non_empty_index < 0:
            return []
        return rows[: last_non_empty_index + 1]
